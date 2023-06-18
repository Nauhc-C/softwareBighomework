import openpyxl

# 创建一个新的workbook和worksheet对象
workbook = openpyxl.Workbook()
worksheet = workbook.active

# 设置第一行标题
worksheet['A1'] = '时间'
worksheet['B1'] = '快充1'
worksheet['H1'] = '快充2'
worksheet['N1'] = '慢充1'
worksheet['T1'] = '慢充2'
worksheet['Z1'] = '慢充3'
# 合并单元格并居中对齐
worksheet.merge_cells('A1:A2')
worksheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet.merge_cells('B1:G1')
worksheet['B1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet.merge_cells('H1:M1')
worksheet['H1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet.merge_cells('N1:S1')
worksheet['N1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet.merge_cells('T1:Y1')
worksheet['T1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet.merge_cells('Z1:AE1')
worksheet['Z1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
#累计充电次数、累计充电时长、累计充电量、累计充电费用、累计服务费用、累计总费用
# 设置第二行标题
for i in range(6):  # 循环4次，每次处理一组标题
    # 计算每组标题的起始列和结束列
    start_col = i * 6 + 2
    end_col = (i + 1) * 6 + 1

    # 设置第二行标题
    worksheet.cell(row=2, column=start_col, value='累计充电次数')
    worksheet.cell(row=2, column=start_col+1, value='累计充电时长')
    worksheet.cell(row=2, column=start_col+2, value='累计充电量')
    worksheet.cell(row=2, column=start_col+3, value='累计充电费用')
    worksheet.cell(row=2, column=start_col+4, value='累计服务费用')
    worksheet.cell(row=2, column=end_col, value='累计总费用')
'''
worksheet['B2'] = '累计充电次数'
worksheet['C2'] = '累计充电时长'
worksheet['D2'] = '累计充电量'
worksheet['E2'] = '累计充电费用'
worksheet['F2'] = '累计服务费用'
worksheet['G2'] = '累计总费用'
'''
# 居中对齐
for cell in ['B2', 'C2']:
    worksheet[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

# 设置宽度
for column_cells in worksheet.columns:
    if column_cells[0].value is None:
        continue  # 跳过合并单元格

    length = max(len(str(cell.value)) for cell in column_cells)
    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 1 if length > 10 else 10

# 写入数据
data = [('China', 'Beijing', 21542000, 'Shanghai', 24256800),
        ('India', 'New Delhi', 18980000, 'Mumbai', 18978000),
        ('USA', 'Washington, D.C.', 633427, 'New York City', 8398748)]

for row in range(3, 6):
    for col in range(1, 5):
        worksheet.cell(row=row, column=col, value=data[row-3][col-1])

# 保存Excel文件
workbook.save('output.xlsx')