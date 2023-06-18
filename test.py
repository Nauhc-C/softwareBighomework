import datetime
import time
import openpyxl
from charge import myTime


def test_create(a):
    time.sleep(1)
    print(a.submit_a_charging_request("2", 100, "T"))
    a.PRINT()
    time.sleep(1)
    print(a.submit_a_charging_request("12332", 100, "F"))
    a.PRINT()
    time.sleep(1)
    print(a.start_charge(6))
    print(a.submit_a_charging_request("asdasd", 100, "T"))
    a.PRINT()
    time.sleep(1)
    print(a.submit_a_charging_request(1, 100, "F"))
    a.PRINT()
    time.sleep(1)
    print(a.submit_a_charging_request(4, 100, "T"))
    time.sleep(1)
    print(a.submit_a_charging_request(11, 100, "T"))
    time.sleep(1)
    print(a.submit_a_charging_request(13, 100, "T"))
    time.sleep(1)
    print(a.submit_a_charging_request(17, 100, "T"))
    time.sleep(1)
    print(a.submit_a_charging_request(19, 100, "T"))
    print(a.waiting_area)

    #print(a.check_pile_report(1))

def test_over(a):
    a.submit_a_charging_request("A", 0.016, "T")
    time.sleep(3)
    a.start_charge("A")
    time.sleep(10)
    print(a.if_car_in_charging("A"))
    print(a.view_billing("A"))

def test_cancel(a):
    a.submit_a_charging_request("A", 0.016, "T")
    time.sleep(3)
    a.end_charge("A")
    a.start_charge("A")
    time.sleep(3)
    time.sleep(10)

def test_pile_open_or_close(a):
    time.sleep(2)
    a.close_pile(2)
    time.sleep(2)
    a.close_pile(2)
    time.sleep(2)
    a.open_pile(2)
    time.sleep(2)

# 测试充电桩数量的
def test_pile_num(a):
    time.sleep(2)
    print(a.retPipeAmount())



def test_pile_broke(a):
    time.sleep(2)
    a.set_pile_error(2)
    time.sleep(2)
    a.set_pile_not_error(2)
    time.sleep(2)


def test_mofity_fee(a):
    time.sleep(3)
    a.submit_a_charging_request("A", 0.5, "T")
    time.sleep(2)
    a.start_charge("A")
    time.sleep(3)
    print(a.view_billing("A"))
    a.setPrice(1000000,2000000,3000000)
    time.sleep(3)
    print(a.view_billing("A"))
# 测试同时给和一段时间内的值
def test_big(a):
    time.sleep(1)
    a.submit_a_charging_request("A", 0.016, "T")
    time.sleep(0.02)
    a.submit_a_charging_request("123", 0.016, "T")
    time.sleep(3)
    print(a.pile_pool[2].waiting_list[0].order_state)
    print(a.look_query("A"))
    a.start_charge("A")
    a.start_charge("123")
    time.sleep(3)
    time.sleep(2)
    print(a.queryReport(2, 2, 9))

def test_last_year(a):
    a.submit_a_charging_request("V1",40,"T")
    time.sleep(1)
    a.start_charge("V1")
    a.submit_a_charging_request("V2",30,"T")
    time.sleep(1)
    a.start_charge("V2")
    a.submit_a_charging_request("V3", 100, "F")
    time.sleep(1)
    a.start_charge("V3")
    a.submit_a_charging_request("V4", 120, "F")
    time.sleep(1)
    a.end_charge("V2")
    time.sleep(1)
    a.start_charge("V4")
    a.submit_a_charging_request("V5", 20, "T")
    time.sleep(1)
    a.start_charge("V5")
    a.submit_a_charging_request("V6", 20, "T")
    time.sleep(1)
    a.start_charge("V6")
    a.submit_a_charging_request("V7", 110, "F")
    time.sleep(1)
    a.start_charge("V7")
    a.submit_a_charging_request("V8", 20, "T")
    time.sleep(1)
    a.start_charge("V8")
    a.set_pile_error(3)
    a.submit_a_charging_request("V9",105,"F")
    time.sleep(1)
    a.start_charge("V9")
    a.submit_a_charging_request("V10",10,"T")
    time.sleep(1)
    a.start_charge("V10")
    a.set_pile_error(0)
    a.submit_a_charging_request("V11",100,"F")
    time.sleep(1)
    a.start_charge("V11")
    a.submit_a_charging_request("V12",90,"F")
    time.sleep(1)
    a.start_charge("V12")
    a.submit_a_charging_request("V13",110,"F")
    time.sleep(1)

    a.start_charge("V13")
    a.submit_a_charging_request("V14",95,"F")

def test_if_car_in_charging(a):
    a.submit_a_charging_request("123123", 16, "T")
    time.sleep(3)
    print("test")
    a.start_charge("123123")
    time.sleep(3)
    print("test")
    print(a.if_car_in_charging("123123"))
    time.sleep(10)

def test_broke_car_start(a):
    a.submit_a_charging_request("qww2",16,"T")
    time.sleep(3)
    a.start_charge("qww2")
    time.sleep(1)
    a.set_pile_error(2)
    time.sleep(1)
    print(a.look_query("qww2"))


def test_start_after_submit(a,name,amount,type):
    a.submit_a_charging_request(name,amount,type)
    time.sleep(3)
    a.start_charge(name)
    time.sleep(2)

def get_list(a,now):
    data=[]
    data.append(now.strftime('%Y-%m-%d %H:%M:%S'))
    data.extend(a.return_pile_information())
    return data


def test_for_big(a):
    # 前期准备工作
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet['A1'] = '时间'
    worksheet['B1'] = '快充1'
    worksheet['H1'] = '快充2'
    worksheet['N1'] = '慢充1'
    worksheet['T1'] = '慢充2'
    worksheet['Z1'] = '慢充3'
    worksheet.merge_cells('A1:A2')
    worksheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.merge_cells('B1:G1')
    worksheet['B1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.merge_cells('H1:M1')
    worksheet['H1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.merge_cells('N1:S1')
    worksheet['N1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.merge_cells('T1:Y1')
    worksheet['T1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.merge_cells('Z1:AE1')
    worksheet['Z1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    for i in range(5):
        start_col = i * 6 + 2
        end_col = (i + 1) * 6 + 1
        worksheet.cell(row=2, column=start_col, value='累计充电次数')
        worksheet.cell(row=2, column=start_col + 1, value='累计充电时长')
        worksheet.cell(row=2, column=start_col + 2, value='累计充电量')
        worksheet.cell(row=2, column=start_col + 3, value='累计充电费用')
        worksheet.cell(row=2, column=start_col + 4, value='累计服务费用')
        worksheet.cell(row=2, column=end_col, value='累计总费用')
    for cell in ['B2', 'C2']:
        worksheet[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    for column_cells in worksheet.columns:
        if column_cells[0].value is None:
            continue

        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = length + 1 if length > 10 else 10








    while True:

        now = myTime.getDataTime()
        #print(now)
        if(now.hour == 5 and now.minute == 43 and now.second==0):
            test_start_after_submit(a,"V1",7,"T")
            test_start_after_submit(a,"V2", 30, "F")  #提前提交


        if(now.hour == 6 and now.minute == 0 and now.second==0):
            print("当前是6点")
            worksheet.append(get_list(a, now))

        if(now.hour == 6 and now.minute == 25 and now.second==0):
            print("当前是6点25")
            test_start_after_submit(a,"V3", 28, "T")
            test_start_after_submit(a,"V4", 120, "F")


        if(now.hour == 6 and now.minute == 30 and now.second==0):
            print("当前是6点30")
            worksheet.append(get_list(a, now))

            a.PRINT()
        if(now.hour == 7 and now.minute == 0 and now.second==0):
            print("当前是7点")
            worksheet.append(get_list(a, now))
            test_start_after_submit(a,"V5", 24.5, "T")
            test_start_after_submit(a,"V6", 45, "F")
            a.PRINT()

        if(now.hour == 7 and now.minute == 30 and now.second==0):
            print("当前是7点30")
            worksheet.append(get_list(a, now))
            a.PRINT()
        if(now.hour == 8 and now.minute == 0 and now.second==0):
            print("当前是8点")
            worksheet.append(get_list(a, now))
            test_start_after_submit(a,"V7", 75, "F")
            a.PRINT()
        if(now.hour == 8 and now.minute == 30 and now.second==0):
            print("当前是8点30")
            worksheet.append(get_list(a, now))
            a.PRINT()
        if(now.hour == 9 and now.minute == 0 and now.second==0):
            print("当前是9点")
            worksheet.append(get_list(a, now))
            test_start_after_submit(a,"V8", 14, "T")
            a.start_charge("V7")
            a.PRINT()
        if(now.hour == 9 and now.minute == 30 and now.second==0):
            print("当前是9点30")
            worksheet.append(get_list(a, now))
            a.PRINT()
        if(now.hour == 10 and now.minute == 0 and now.second==0):
            print("当前是10点")
            worksheet.append(get_list(a, now))
            a.set_pile_error(4)
            a.PRINT()
        if(now.hour == 10 and now.minute == 30 and now.second==0):
            print("当前是10点30")
            worksheet.append(get_list(a, now))
            a.PRINT()
        if(now.hour == 11 and now.minute == 0 and now.second==0):
            print("当前是11点")
            worksheet.append(get_list(a, now))
            a.PRINT()
        if(now.hour == 11 and now.minute == 30 and now.second==0):
            print("当前是11点30")
            worksheet.append(get_list(a, now))
            a.PRINT()
        if(now.hour == 12 and now.minute == 0 and now.second==0):
            print("当前是12点")
            worksheet.append(get_list(a, now))
            a.PRINT()
            break
        time.sleep(1)

    workbook.save('output.xlsx')
